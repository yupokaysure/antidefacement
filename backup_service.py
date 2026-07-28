from __future__ import annotations

import asyncio
import json
import logging
import secrets
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import discord
from openpyxl import Workbook
from openpyxl.styles import Font

from config import BACKUP_DIR, DEFAULT_TIMEZONE
from serializers import serialize_channel, serialize_member, serialize_role
from storage import PostgresStore, utc_iso

log = logging.getLogger(__name__)


class BackupService:
    def __init__(self, bot: discord.Client, store: PostgresStore) -> None:
        self.bot = bot
        self.store = store

    async def generate_backup_id(self) -> str:
        while True:
            candidate = f"BK-{datetime.now(timezone.utc):%Y%m%d}-{secrets.randbelow(10**10):010d}"
            if await self.store.get_backup(candidate) is None:
                return candidate

    async def build_snapshot(self, guild: discord.Guild) -> dict[str, Any]:
        try:
            if guild.large and not guild.chunked:
                await guild.chunk(cache=True)
        except (discord.HTTPException, discord.ClientException):
            log.warning("Could not chunk guild %s before backup", guild.id)

        return {
            "schema_version": 1,
            "guild": {
                "id": guild.id,
                "name": guild.name,
                "owner_id": guild.owner_id,
                "member_count": guild.member_count,
                "created_at": guild.created_at.isoformat(),
                "backed_up_at": utc_iso(),
            },
            "roles": [serialize_role(role) for role in guild.roles],
            "channels": [serialize_channel(channel) for channel in guild.channels],
            "members": [serialize_member(member) for member in guild.members],
        }

    async def create_backup(
        self,
        guild: discord.Guild,
        *,
        initiated_by: int | None,
        backup_type: str,
        deliver: bool = True,
    ) -> dict[str, Any]:
        backup_id = await self.generate_backup_id()
        metadata = {
            "backup_id": backup_id,
            "guild_id": guild.id,
            "guild_name": guild.name,
            "created_at": utc_iso(),
            "created_by": initiated_by,
            "backup_type": backup_type,
            "status": "creating",
            "snapshot": None,
            "delivery_status": None,
            "delivery_channel_id": None,
            "delivery_message_id": None,
            "error": None,
        }
        await self.store.add_backup(metadata)

        try:
            snapshot = await self.build_snapshot(guild)
            await self.store.update_backup(
                backup_id,
                status="complete",
                snapshot=snapshot,
            )
        except Exception as exc:
            log.exception("Backup %s failed", backup_id)
            await self.store.update_backup(backup_id, status="failed", error=str(exc))
            raise

        backup = (await self.store.get_backup(backup_id)) or metadata
        if deliver:
            delivery = await self.deliver_backup(guild, backup)
            await self.store.update_backup(backup_id, **delivery)

        return (await self.store.get_backup(backup_id)) or backup

    async def export_backup_files(self, backup: dict[str, Any]) -> list[str]:
        snapshot = await self.load_snapshot(backup)
        guild_dir = BACKUP_DIR / str(backup["guild_id"])
        guild_dir.mkdir(parents=True, exist_ok=True)
        json_path = guild_dir / f"{backup['backup_id']}.json"
        xlsx_path = guild_dir / f"{backup['backup_id']}.xlsx"
        await asyncio.to_thread(
            json_path.write_text,
            json.dumps(snapshot, indent=2, ensure_ascii=False),
            "utf-8",
        )
        await asyncio.to_thread(self._write_workbook, snapshot, xlsx_path)
        return [str(json_path), str(xlsx_path)]

    @staticmethod
    async def cleanup_export_files(paths: list[str]) -> None:
        for value in paths:
            try:
                await asyncio.to_thread(Path(value).unlink, missing_ok=True)
            except OSError:
                pass

    async def deliver_backup(self, guild: discord.Guild, backup: dict[str, Any]) -> dict[str, Any]:
        settings = await self.store.get_guild(guild.id)
        channel_id = settings.get("backup_channel_id")
        channel = guild.get_channel(int(channel_id)) if channel_id else None
        if not isinstance(channel, discord.abc.Messageable):
            return {
                "delivery_status": "not_configured",
                "delivery_channel_id": channel_id,
                "delivery_message_id": None,
            }

        paths: list[str] = []
        try:
            paths = await self.export_backup_files(backup)
            message = await channel.send(
                content=(
                    f"Anti-Defacement backup **{backup['backup_id']}** for **{guild.name}**. "
                    "PostgreSQL is the restoration source; the JSON and spreadsheet are exports."
                ),
                files=[discord.File(value, filename=Path(value).name) for value in paths],
            )
            return {
                "delivery_status": "sent",
                "delivery_channel_id": channel.id,
                "delivery_message_id": message.id,
            }
        except (discord.Forbidden, discord.HTTPException, OSError) as exc:
            log.warning("Could not deliver backup %s: %s", backup["backup_id"], exc)
            return {
                "delivery_status": f"failed: {exc}",
                "delivery_channel_id": channel.id,
                "delivery_message_id": None,
            }
        finally:
            await self.cleanup_export_files(paths)

    @staticmethod
    def _style_header(sheet) -> None:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    def _write_workbook(self, snapshot: dict[str, Any], path: Path) -> None:
        workbook = Workbook()
        members_sheet = workbook.active
        members_sheet.title = "Members"
        members_sheet.append(
            [
                "Discord ID",
                "Username",
                "Display Name",
                "Nickname",
                "Bot",
                "Joined At",
                "Created At",
                "Role IDs",
                "Role Names",
            ]
        )
        for member in snapshot["members"]:
            members_sheet.append(
                [
                    str(member.get("id", "")),
                    member.get("username", ""),
                    member.get("display_name", ""),
                    member.get("nickname", ""),
                    member.get("bot", False),
                    member.get("joined_at", ""),
                    member.get("created_at", ""),
                    ", ".join(str(value) for value in member.get("role_ids", [])),
                    ", ".join(member.get("role_names", [])),
                ]
            )
        self._style_header(members_sheet)

        role_sheet = workbook.create_sheet("Roles")
        role_sheet.append(
            [
                "Role ID",
                "Name",
                "Position",
                "Color Integer",
                "Color Hex",
                "Permissions Integer",
                "Hoisted",
                "Mentionable",
                "Managed",
                "Default Role",
                "Unicode Emoji",
            ]
        )
        for role in snapshot["roles"]:
            color = int(role.get("color", 0))
            role_sheet.append(
                [
                    str(role.get("id", "")),
                    role.get("name", ""),
                    role.get("position", 0),
                    color,
                    f"#{color:06X}",
                    str(role.get("permissions", 0)),
                    role.get("hoist", False),
                    role.get("mentionable", False),
                    role.get("managed", False),
                    role.get("is_default", False),
                    role.get("unicode_emoji", ""),
                ]
            )
        self._style_header(role_sheet)

        channel_sheet = workbook.create_sheet("Channels")
        channel_sheet.append(
            [
                "Channel ID",
                "Name",
                "Type",
                "Position",
                "Category ID",
                "Category Name",
                "Topic",
                "NSFW",
                "Slowmode",
                "Bitrate",
                "User Limit",
            ]
        )
        for channel in snapshot["channels"]:
            channel_sheet.append(
                [
                    str(channel.get("id", "")),
                    channel.get("name", ""),
                    channel.get("kind", ""),
                    channel.get("position", 0),
                    str(channel.get("category_id") or ""),
                    channel.get("category_name", ""),
                    channel.get("topic", ""),
                    channel.get("nsfw", False),
                    channel.get("slowmode_delay", 0),
                    channel.get("bitrate", ""),
                    channel.get("user_limit", ""),
                ]
            )
        self._style_header(channel_sheet)

        overwrite_sheet = workbook.create_sheet("Channel Overwrites")
        overwrite_sheet.append(
            [
                "Channel ID",
                "Channel Name",
                "Target ID",
                "Target Name",
                "Target Type",
                "Allow Integer",
                "Deny Integer",
            ]
        )
        for channel in snapshot["channels"]:
            for overwrite in channel.get("overwrites", []):
                overwrite_sheet.append(
                    [
                        str(channel.get("id", "")),
                        channel.get("name", ""),
                        str(overwrite.get("target_id", "")),
                        overwrite.get("target_name", ""),
                        overwrite.get("target_type", ""),
                        str(overwrite.get("allow", 0)),
                        str(overwrite.get("deny", 0)),
                    ]
                )
        self._style_header(overwrite_sheet)

        for sheet in workbook.worksheets:
            for column in sheet.columns:
                width = min(max((len(str(cell.value or "")) for cell in column), default=8) + 2, 60)
                sheet.column_dimensions[column[0].column_letter].width = width

        workbook.save(path)

    @staticmethod
    def compute_next_run(schedule: dict[str, Any], *, now: datetime | None = None) -> datetime:
        timezone_name = schedule.get("timezone") or DEFAULT_TIMEZONE
        try:
            zone = ZoneInfo(timezone_name)
        except ZoneInfoNotFoundError as exc:
            raise ValueError(f"Unknown timezone: {timezone_name}") from exc

        now_utc = now or datetime.now(timezone.utc)
        local_now = now_utc.astimezone(zone)
        hour = int(schedule.get("hour", 3))
        minute = int(schedule.get("minute", 0))
        frequency = schedule.get("frequency", "daily")

        candidate = local_now.replace(hour=hour, minute=minute, second=0, microsecond=0)

        if frequency == "daily":
            if candidate <= local_now:
                candidate += timedelta(days=1)
        elif frequency == "weekly":
            weekday = int(schedule.get("weekday", 0))
            days_ahead = (weekday - local_now.weekday()) % 7
            candidate += timedelta(days=days_ahead)
            if candidate <= local_now:
                candidate += timedelta(days=7)
        elif frequency == "monthly":
            day = max(1, min(int(schedule.get("day_of_month", 1)), 28))
            candidate = candidate.replace(day=day)
            if candidate <= local_now:
                if candidate.month == 12:
                    candidate = candidate.replace(year=candidate.year + 1, month=1)
                else:
                    candidate = candidate.replace(month=candidate.month + 1)
        else:
            raise ValueError("Frequency must be daily, weekly, or monthly")

        return candidate.astimezone(timezone.utc)

    async def load_snapshot(self, backup: dict[str, Any]) -> dict[str, Any]:
        snapshot = backup.get("snapshot")
        if not isinstance(snapshot, dict):
            fresh = await self.store.get_backup(str(backup["backup_id"]))
            snapshot = fresh.get("snapshot") if fresh else None
        if not isinstance(snapshot, dict):
            raise RuntimeError(
                f"Backup {backup.get('backup_id')} has no PostgreSQL snapshot."
            )
        return snapshot
